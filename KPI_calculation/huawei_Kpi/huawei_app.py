import xml.etree.ElementTree as ET
import re 
from openpyxl import Workbook, load_workbook, utils
import os 


# tree = ET.parse('/home/swapan/dev/KPI_calculation/Src/huawei_kpi.xml')

def parse_input_xml(xml_file_path, mand_kpi:set, input_dict: dict):
    mand_kpi=mand_kpi
    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    print("root tag is {}".format(root.tag))
    prefix_pattern = re.compile(r"{.*}")
    tag_prefix = re.findall(prefix_pattern, root.tag)[0]
    print("tag-prefix is {}".format(tag_prefix))
    begin_time = root.find("{0}fileHeader/{0}measCollec".format(tag_prefix)).attrib["beginTime"][0:19]
    meas_data_element = root.find(tag_prefix+"measData")
    print("meas_data_element is " + meas_data_element.tag)
    measInfo_pattrn = re.compile('measInfo')
    meas_info_list = [chld_elements  for chld_elements in meas_data_element if re.search(measInfo_pattrn, chld_elements.tag)]
    # Under each meas_info i have multiple child elements having actual data 
    # What I expect from the loop below, for each meas_id all the results
    input_dict = input_dict 
    for meas_info in meas_info_list:
        meas_info_id = meas_info.attrib.get('measInfoId')
        duration = meas_info.find(tag_prefix + "granPeriod").attrib.get('duration')[2:-1]
        Start_time_duration_meas_info_id = "{}-{}-{}".format(begin_time, duration, meas_info_id)
        meas_type = meas_info.find(tag_prefix +   "measTypes")
        meas_types = str(meas_type.text).split()
        # What I am expecting from the loop below, a dict of node_id vs related results/value for kpi
        node_v_meas_result = dict()
        for element in meas_info.findall(tag_prefix+"measValue"):
            node_id = None
            meas_result = None
            meas_results = None
            node_id = element.attrib["measObjLdn"]
            meas_result = list(element)[0]
            meas_results = str(meas_result.text).split()
            meas_type_v_result = dict(zip(meas_types, meas_results))
            for meas_type in meas_types:
                if int(meas_type) not in mand_kpi:
                    meas_type_v_result.pop(meas_type)
            node_v_meas_result[node_id] = meas_type_v_result

        if input_dict.get(str(Start_time_duration_meas_info_id)) is None:
            # This to handle if I get multiple entry of same measinfoid for a beging-time and duration 
            input_dict[str(Start_time_duration_meas_info_id)] = []
            input_dict[str(Start_time_duration_meas_info_id)].append(node_v_meas_result)
        else:
            input_dict[str(Start_time_duration_meas_info_id)].append(node_v_meas_result)

    return input_dict


def read_mandatory_kpi(filter_file):
    kpi_filter = load_workbook(filter_file)
    counter_sheet = kpi_filter['Counters']
    rows = counter_sheet.rows
    headers_row = next(rows)
    # headers_row = counter_sheet[1]
    kpi_column = None
    for header_cell in headers_row:
        if header_cell.value.strip() == "KPI_Number":
            kpi_column = utils.cell.get_column_letter(header_cell.column)
            break
    kpi_numbers = [cell.value for cell in counter_sheet[kpi_column] if cell.row != 1]
    mandatory_kpi = set(kpi_numbers)
    print(mandatory_kpi)
    # mandatory_kpi = set([50332573,50332574,50342574,50342575,50342635,50342636,50332582,50332583,50332584,50332586,50342584,50342585,50342586,50342587,50342606,50342607,50342608,50342609,50342632,50342633,50332582,50332583,50332584,50332586,50342584,50342585,50342586,50342587,50342606,50342607,50342608,50342609,50342632,50342633,50332582,50332583,50332584,50332586,50342584,50342585,50342586,50342587,50342606,50342607,50342608,50342609,50342632,50342633,50332627,50332627,50332627,1526727075,1526727076,1526727077,1526727078,1526727079,1526727080,1542455297,1542455298,1542455299,1542455300,1542455301,1542455302,1542455303,1542455304,1542455305,1542455306,1542455307,1542460296,1542460297,1542455297,1542455298,1542455299,1542455300])
    return mandatory_kpi


def write_to_excel(dict_to_be_stored, output_dir):
    input_dict = dict_to_be_stored
    output_file_path = os.path.join (output_dir, "output.xlsx")
    Wb = Workbook()
    sheet = Wb.worksheets[0]
    sheet.title = "Huawei_KPI"
    headers = enumerate(["DateTime(UTC)-Duration-Meas_id", "Node_id"], start=1)
    for index, value in headers:
        sheet.cell(row=1, column=index, value=value)

    # for row_id, meas_id in enumerate(input_dict, start=2):
    #     # start = 2, as first row is header
    #     sheet.cell(row_id, 1, int(meas_id))
    #     # I don't want to proceed next row before I write other columns
    #     for row_id_dlt1, node_id in input_dict[meas_id]
    # input_dict1 = {'50331655': [{'MBTSCHQ045/ULGROUP:UL BB Resource Group No.=0': {'50332573': '384', '50332574': '186.038', '50342574': '160', '50342575': '239.5', '50342635': '0', '50342636': '0'}}], '50331656': [{'MBTSCHQ045/BRD:Board Type=BBP, Cabinet No.=0, Subrack No.=0, Slot No.=3': {'50332582': '512', '50332583': '384', '50332584': '94.542', '50332586': '186.038', '50342584': '83', '50342585': '109', '50342586': '160', '50342587': '239.5', '50342606': '27', '50342607': '35', '50342608': '29', '50342609': '36', '50342632': '0', '50342633': '0'}}]}
    meas_type_history = set()
    col_id = (3+len(meas_type_history))
    row_ind = 2
    # for row_ind, meas_id in enumerate(input_dict, start=2):
    for meas_id in input_dict:
        meas_values = input_dict[meas_id][0]
        final_row_id = 0 # Need to store this value while come out from the loop below
        for row_ind_chld,  node_id in enumerate(meas_values.keys()):
            meas_type_v_result = meas_values[node_id] # result is a dict
            final_row_id = (row_ind+row_ind_chld)
            # TODO Debug print
            # print("{}={}".format(final_row_id, meas_type_v_result))

            for meas_type in meas_type_v_result:
                # print("column_id is {}".format(col_id))
                sheet.cell(final_row_id, column=1, value=meas_id)
                sheet.cell(final_row_id, column=2, value=node_id)
                # check if meas_type already there and get it cell location
                if meas_type in meas_type_history:
                    # print("{} exist ".format(meas_type))
                    col_id_exist = None # Need to get column where meas_type is populated
                    for i in range(3, sheet.max_column+1):
                        if sheet.cell(1, i).value == meas_type:
                            col_id_exist = i
                            break
                        else:
                            continue
                    # sheet.cell(1, col_id_exist, value=meas_type) # this value is already populated
                    if meas_type_v_result[meas_type] == "NIL":
                        sheet.cell(final_row_id, col_id_exist, 0)
                    else:
                        sheet.cell(final_row_id, col_id_exist, float(meas_type_v_result[meas_type]))
                else:
                    meas_type_history.add(meas_type)
                    sheet.cell(1, col_id, value=meas_type)
                    if meas_type_v_result[meas_type] == "NIL":
                        sheet.cell(final_row_id, col_id, 0)
                        col_id = (col_id+1)
                    else:
                        sheet.cell(final_row_id, col_id, float(meas_type_v_result[meas_type]))
                        col_id = (col_id + 1)
        row_ind = (final_row_id + 1)


    try:
        Wb.save(output_file_path)
    except:
        print("Error while saving")
    else:
        print("**********Writing done and saved ***************")


if __name__ == "__main__":    
    # mandatory_kpi = set([67179340,67179339,67189465,67189468,67189470,67189466,67189467,67189472,67189471,67189459,67189462,67189469,67189460,67189461,67189464,67189463,67179299,67192570,67192571,67180378,67199736,73410510,67189768,67183489,67183490,67189832,67189833,67202894,67189840,67202984,73410492,67191162,67191163,67191164,67191161,67190705,73403763,67202932,73410491,67203932,67192486,67192115,73403761,67203850,73410507,67183495,67183497,67189908,67189907,67191155,67191156,67199682,67202900,67199617,67199618,67202902,67199681,67199683,67202901,67202943,67199692,67202944,67199693,67202942,67199694,67199691,67183997,67183998,67183994,67183999,67183995,67184000,67183996,67183993,67184018,67184019,67184015,67184020,67184016,67184021,67184017,67184014,67184005,67184006,67184002,67184007,67184003,67184008,67184004,67184001,67184026,67184027,67184023,67184028,67184024,67184029,67184025,67184022,67193548,67193550,67184009,67184010,67184011,67184012,67193552,67193554,67193556,67184030,67184031,67184032,67193558,73393910,73393908,67180082,67179778,67179777,67180077,67180083,67179781,67179780,67174955,67180078,67180079,67174964,67179858,67179825,67179826,67179924,67179921,67179923,67179922,67179830,67179930,67190518,67191786,67179782,67174966,67190455,67179860,67179827,67179828,67179928,67179925,67179927,67179926,67199619,67199620,67199625,67199624,67199623,67199780,67203809,67203810,67203819,67199556,67190586,67199510,67179521,67190587,67203936,67203937,67203938,67189910,67180498,67180508,67189909,67180499,67180509,67199703,67199700,67199704,67199701,67199698,67199702,67199699,67180483,67192572,67189767])
    # mandatory_kpi1 = set([50332573,50332574,50342574,50342575,50342635,50342636,50332582,50332583,50332584,50332586,50342584,50342585,50342586,50342587,50342606,50342607,50342608,50342609,50342632,50342633,50332582,50332583,50332584,50332586,50342584,50342585,50342586,50342587,50342606,50342607,50342608,50342609,50342632,50342633,50332582,50332583,50332584,50332586,50342584,50342585,50342586,50342587,50342606,50342607,50342608,50342609,50342632,50342633,50332627,50332627,50332627,1526727075,1526727076,1526727077,1526727078,1526727079,1526727080,1542455297,1542455298,1542455299,1542455300,1542455301,1542455302,1542455303,1542455304,1542455305,1542455306,1542455307,1542460296,1542460297,1542455297,1542455298,1542455299,1542455300])
    #
    # kpi_flie = r"D:\D_drive_BACKUP\study\PycharmProjects\KPI_calculation\KPI_calculation\huawei_kpi.xml"
    # ouput_dir = r"D:\D_drive_BACKUP\study\PycharmProjects\KPI_calculation\output_folder"
    # input_dict = parse_input_xml(kpi_flie, mandatory_kpi1)
    # print(input_dict)
    # print("**********Writing into excel file***************")
    # """"""
    # write_to_excel(input_dict, ouput_dir)
    # print(mandatory_kpi)

    read_mandatory_kpi(r"D:\D_drive_BACKUP\study\PycharmProjects\KPI_calculation\Data\filter\Huawei_LTE_filter.xlsx")
